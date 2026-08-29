from __future__ import annotations

import pytest

from silica.tools.tabular import silica_query_table


@pytest.fixture
def sales(tmp_path):
    p = tmp_path / "sales.csv"
    p.write_text(
        "region,amount,year\n"
        "north,100,2024\nnorth,50,2025\nsouth,30,2024\nsouth,20,2024\n"
    )
    return p


def test_aggregates(sales):
    out = silica_query_table(
        str(sales), "SELECT region, sum(amount) AS tot FROM t GROUP BY 1 ORDER BY 1"
    )
    assert out["columns"] == ["region", "tot"]
    assert out["rows"] == [["north", 150], ["south", 50]]
    assert out["truncated"] is False


@pytest.mark.parametrize(
    "sql",
    [
        "DROP TABLE t",
        "COPY (SELECT 1) TO '/tmp/pwn.csv'",
        "ATTACH '/tmp/x.db'",
        "SELECT 1; DROP TABLE t",  # the second statement is the payload
    ],
)
def test_rejects_non_select(sales, sql):
    with pytest.raises(ValueError):
        silica_query_table(str(sales), sql)


def test_confined_to_the_targets_directory(sales, tmp_path):
    """A SELECT is not enough: read_csv() inside one can still name any path."""
    outside = tmp_path.parent / "secret.csv"
    outside.write_text("s\n42\n")
    with pytest.raises(ValueError, match="query failed"):
        silica_query_table(str(sales), f"SELECT * FROM read_csv('{outside}')")


def test_truncation_is_flagged_not_silent(sales):
    out = silica_query_table(str(sales), "SELECT * FROM t", limit=2)
    assert out["row_count"] == 2 and out["truncated"] is True
    assert "limit=2" in out["note"]


def test_rejects_non_tabular_extension(tmp_path):
    note = tmp_path / "note.md"
    note.write_text("# hi")
    with pytest.raises(ValueError, match="not a tabular format"):
        silica_query_table(str(note), "SELECT * FROM t")


# --- schema and type honesty -------------------------------------------------

def test_schema_rides_on_every_reply(sales):
    """SQL written against guessed column types is the lane's failure mode."""
    out = silica_query_table(str(sales), "SELECT 1 AS one FROM t LIMIT 1")
    assert out["schema"] == {"region": "VARCHAR", "amount": "BIGINT", "year": "BIGINT"}


def test_full_file_sniff_survives_a_late_string(tmp_path):
    """DuckDB's default sample types on the head; row 21001 must not crash."""
    p = tmp_path / "long.csv"
    p.write_text("v\n" + "1\n" * 21000 + "x\n")
    out = silica_query_table(str(p), "SELECT count(*) AS n FROM t")
    assert out["rows"] == [[21001]]
    assert out["schema"]["v"] == "VARCHAR"  # typed on the whole file, not the head


def test_varchar_aggregate_fails_loud_not_wrong(tmp_path):
    """European numerics + n/a: sum() must error, never a partial number."""
    p = tmp_path / "eu.csv"
    p.write_text('city,revenue\nRoma,"1.234,50"\nMilano,987\nRoma,n/a\n')
    out = silica_query_table(str(p), "SUMMARIZE t")  # the taught first call
    assert out["schema"]["revenue"] == "VARCHAR"
    with pytest.raises(ValueError, match="VARCHAR"):
        silica_query_table(str(p), "SELECT sum(revenue) FROM t")


def test_byte_cap_truncates_wide_payloads(tmp_path):
    """`limit` counts rows; the cap counts bytes. Both must announce themselves."""
    p = tmp_path / "wide.csv"
    p.write_text("c\n" + ("x" * 1000 + "\n") * 200)
    out = silica_query_table(str(p), "SELECT * FROM t")
    assert out["truncated"] is True
    assert 0 < out["row_count"] < 200
    assert "KB" in out["note"]


# --- Excel: one sheet per call, via the temp-CSV detour ----------------------

def _xlsx(tmp_path, sheets: dict) -> str:
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    p = tmp_path / "book.xlsx"
    wb.save(p)
    return str(p)


def test_xlsx_single_sheet_needs_no_sheet_arg(tmp_path):
    import datetime

    p = _xlsx(
        tmp_path,
        {"Data": [
            ["day", "amount"],
            [datetime.date(2025, 1, 1), 100],
            [datetime.date(2025, 2, 1), 50],
        ]},
    )
    out = silica_query_table(p, "SELECT sum(amount) AS tot FROM t")
    assert out["rows"] == [[150]]
    assert out["sheet"] == "Data"
    # a date cell survives the CSV round-trip as a date, not as text
    assert out["schema"]["day"] in ("DATE", "TIMESTAMP")


def test_xlsx_multi_sheet_rejects_and_lists(tmp_path):
    p = _xlsx(tmp_path, {"A": [["x"], [1]], "B": [["y"], [2]]})
    with pytest.raises(ValueError, match="A, B"):
        silica_query_table(p, "SELECT * FROM t")
    out = silica_query_table(p, "SELECT * FROM t", sheet="B")
    assert out["columns"] == ["y"]


def test_sheet_arg_on_a_csv_is_rejected(sales):
    with pytest.raises(ValueError, match="Excel"):
        silica_query_table(str(sales), "SELECT * FROM t", sheet="Data")


def test_vault_relative_path_resolves(tmp_path, monkeypatch):
    # The profile note cites the file repo-relative; the example it prints must
    # be copy-pasteable, so the tool tries the vault before the cwd.
    import silica.config

    vault = tmp_path / "v"
    (vault / "data").mkdir(parents=True)
    (vault / "data" / "s.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    monkeypatch.setattr(silica.config.CONFIG, "vault_path", str(vault))
    monkeypatch.chdir(tmp_path)  # cwd resolution alone would miss

    out = silica_query_table("data/s.csv", "SELECT count(*) AS c FROM t")
    assert out["rows"] == [[1]]


def test_latin1_csv_is_readable(tmp_path):
    # ISTAT and most European public exports ship windows-1252/latin-1, not
    # utf-8; DuckDB's reader rejects those outright with a raw parser error.
    p = tmp_path / "legenda.csv"
    p.write_bytes(
        "codice;descrizione\nP2;Variazione pi\xf9 che annua\n".encode("latin-1")
    )
    out = silica_query_table(str(p), "SELECT descrizione FROM t")
    assert out["rows"] == [["Variazione pi\xf9 che annua"]]


# --- silica_tables: the census before the query ---------------------------


@pytest.fixture
def table_vault(tmp_path, monkeypatch):
    import silica.config

    (tmp_path / "data").mkdir()
    (tmp_path / "data" / "sales.csv").write_text(
        "region,amount\nnorth,1\n", encoding="utf-8"
    )
    (tmp_path / "data" / "neet.csv").write_text(
        "comune,neet_rate,year\nbagheria,0.31,2021\n", encoding="utf-8"
    )
    (tmp_path / "note.md").write_text("# not a table", encoding="utf-8")
    (tmp_path / "prose.txt").write_text("plain text stays out", encoding="utf-8")
    (tmp_path / "vendored").mkdir()
    (tmp_path / "vendored" / "junk.csv").write_text("x\n1\n", encoding="utf-8")
    (tmp_path / ".silicaignore").write_text("vendored\n", encoding="utf-8")
    monkeypatch.setattr(silica.config.CONFIG, "vault_path", str(tmp_path))
    return tmp_path


def test_census_lists_schemas_not_prose(table_vault):
    from silica.tools.tabular import silica_tables

    out = silica_tables()
    paths = {t["path"] for t in out["tables"]}
    assert paths == {"data/neet.csv", "data/sales.csv"}
    by_path = {t["path"]: t for t in out["tables"]}
    assert set(by_path["data/neet.csv"]["columns"]) == {"comune", "neet_rate", "year"}
    assert out["total"] == 2 and out["truncated"] is False


def test_census_respects_silicaignore(table_vault):
    from silica.tools.tabular import silica_tables

    assert all("vendored" not in t["path"] for t in silica_tables()["tables"])


def test_column_filter_names_its_matches(table_vault):
    from silica.tools.tabular import silica_tables

    out = silica_tables(column="NEET")
    assert [t["path"] for t in out["tables"]] == ["data/neet.csv"]
    assert out["tables"][0]["matched"] == ["neet_rate"]


def test_census_scopes_to_a_folder(table_vault):
    from silica.tools.tabular import silica_tables

    (table_vault / "other").mkdir()
    (table_vault / "other" / "x.csv").write_text("a\n1\n", encoding="utf-8")
    out = silica_tables(folder="data")
    assert {t["path"] for t in out["tables"]} == {"data/neet.csv", "data/sales.csv"}


def test_census_limit_truncates_with_a_note(table_vault):
    from silica.tools.tabular import silica_tables

    out = silica_tables(limit=1)
    assert out["returned"] == 1 and out["truncated"] is True
    assert "limit=1" in out["note"]
    assert out["total"] == 2  # the census size survives the cap


def test_unreadable_file_is_a_named_entry_not_a_dropped_path(table_vault):
    from silica.tools.tabular import silica_tables

    (table_vault / "data" / "broken.parquet").write_bytes(b"\x00not parquet")
    out = silica_tables()
    broken = [t for t in out["tables"] if t["path"] == "data/broken.parquet"]
    assert broken and "error" in broken[0]


def test_census_escape_is_rejected(table_vault):
    from silica.tools.tabular import silica_tables

    with pytest.raises(ValueError, match="escapes"):
        silica_tables(folder="../outside")


def test_xlsx_entries_list_sheets(table_vault):
    import openpyxl

    from silica.tools.tabular import silica_tables

    wb = openpyxl.Workbook()
    wb.active.title = "Anagrafica"
    wb.create_sheet("Pendolarismo")
    wb.save(table_vault / "data" / "book.xlsx")
    res = silica_tables()
    entry = [t for t in res["tables"] if t["path"] == "data/book.xlsx"]
    assert entry and entry[0]["sheets"] == ["Anagrafica", "Pendolarismo"]
