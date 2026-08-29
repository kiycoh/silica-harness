# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Alessandro Carosia

"""Tabular query tool — read-only SQL over a data file, via DuckDB.

The BI probe lane. Rows cannot survive the SourceAdapter contract (ADR-0014
turns every source into markdown prose), and no amount of rerank tuning makes
"revenue by region" a top-k similarity problem — so tabular data gets its own
retrieval path instead of being forced through recall. This module is that
path's whole surface: one tool, one dependency, no ETL and no server.

Zero-trust (ADR-0009): the SQL is model-authored, so it is parsed and rejected
unless it is a single SELECT, and the connection is confined to the target
file's own directory before the query runs.
"""
from __future__ import annotations

import io
import json
import logging
import tempfile
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field

from silica.tools import tool

logger = logging.getLogger(__name__)

# Everything DuckDB binds from a bare path with no extension loaded.
READABLE_EXTS = (".csv", ".tsv", ".txt", ".parquet", ".json", ".ndjson")
# Spreadsheets: one sheet per call, extracted in-process to a temp CSV that is
# then queried like any other. The direct route (DuckDB's `excel` extension)
# needs INSTALL, which the sandbox below blocks by design — the detour keeps
# the whole downstream path (sniffing, confinement, caps) shared instead of
# duplicated per format.
SHEET_EXTS = (".xlsx", ".xlsm", ".xls")
# Text formats are sniffed on the WHOLE file (sample_size=-1), not DuckDB's
# default 20480-row head: a column typed on its head either errors on the first
# odd row past the sample or — worse — bakes a type the data does not keep.
# Parquet/JSON carry their own types, so the knob does not apply to them.
_SNIFFED_EXTS = (".csv", ".tsv", ".txt")

# Rows are capped in bytes as well as in count: `limit` alone never sees that
# 200 rows of a wide table is a six-figure token payload.
_PAYLOAD_BYTE_CAP = 50_000


def _connect(data_dir: Path):
    """A DuckDB connection that can read `data_dir` and nothing else.

    The order is load-bearing and was verified, not assumed: `allowed_directories`
    on its own confines nothing (it read /etc/passwd and wrote via COPY TO). It
    is an allowlist carved out of `enable_external_access=false`, so the access
    flag must be dropped *after* the directory is named, and the configuration
    locked *after* both — otherwise the model's SQL can simply widen it back.
    """
    import duckdb

    con = duckdb.connect()
    con.execute("SET allowed_directories=[?]", [str(data_dir)])
    con.execute("SET enable_external_access=false")
    con.execute("SET lock_configuration=true")
    return con


def _select_only(sql: str) -> str:
    """The one statement in `sql` if it is a SELECT, else ValueError.

    Parsed by DuckDB, not pattern-matched: a keyword blocklist misses `ATTACH`,
    misreads a column literally named "drop", and is exactly the flimsy half of
    a choice where the correct half costs the same one call.
    """
    import duckdb

    try:
        statements = duckdb.extract_statements(sql)
    except Exception as e:
        raise ValueError(f"unparsable SQL: {e}") from e
    if len(statements) != 1:
        raise ValueError(
            f"expected exactly 1 statement, got {len(statements)} — "
            "this tool reads, so it runs one SELECT per call"
        )
    stmt = statements[0]
    if stmt.type != duckdb.StatementType.SELECT:
        raise ValueError(
            f"{stmt.type.name} rejected: silica_query_table is read-only, "
            "the query must be a SELECT (WITH … SELECT is fine)"
        )
    return stmt.query.strip().rstrip(";")


def _bind_source(con, path: Path) -> None:
    """CREATE VIEW t over `path` — read per query, never copied.

    Interpolated because a path cannot be a prepared-statement parameter in
    FROM; the path is resolved and real, and the quote-doubling closes the
    only injection route left.
    """
    quoted = str(path).replace("'", "''")
    if path.suffix.lower() in _SNIFFED_EXTS:
        con.execute(
            f"CREATE VIEW t AS SELECT * FROM read_csv('{quoted}', sample_size=-1)"
        )
    else:
        con.execute(f"CREATE VIEW t AS SELECT * FROM '{quoted}'")


# Western-European ladder, tried in order. utf-8 first so a correct file is
# never re-encoded; cp1252 before latin-1 because Windows exports put curly
# quotes and the ellipsis in the C1 range (ISTAT's "…" is 0x85, which latin-1
# reads as the NEL control); latin-1 last because it decodes every byte and
# so can never itself raise.
_TEXT_ENCODINGS = ("utf-8", "cp1252", "latin-1")


def utf8_source(src: Path, tmpdir: Path) -> Path:
    """`src` when it is already utf-8, else a re-encoded copy under `tmpdir`.

    DuckDB's reader validates encoding and refuses anything else outright —
    including bytes its own `encoding='latin-1'` rejects (0x85 again) — so the
    decode happens here, in Python, where the ladder is ours to pick. Measured
    on a real public-data vault: 4 of 26 CSVs were not utf-8.
    """
    raw = src.read_bytes()
    for enc in _TEXT_ENCODINGS:
        try:
            text = raw.decode(enc)
        except UnicodeDecodeError:
            continue
        if enc == "utf-8":
            return src            # the common case copies nothing
        out = tmpdir / src.name
        out.write_text(text, encoding="utf-8")
        return out
    return src  # undecodable by latin-1 is impossible; bind and let DuckDB speak


def _pick_sheet(names: list[str], sheet: str) -> str:
    """The requested sheet name, or the only one; ambiguity is a teaching error."""
    if sheet:
        if sheet in names:
            return sheet
        raise ValueError(f"sheet {sheet!r} not found — this workbook has: {', '.join(names)}")
    if len(names) == 1:
        return names[0]
    # Guessing (say, the first sheet) would answer from the wrong table without
    # any signal that it did; the retry costs one call and names the choices.
    raise ValueError(
        "this workbook has several sheets — pass sheet= to pick one of: " + ", ".join(names)
    )


def _sheet_rows(src: Path, sheet: str):
    """(sheet name, iterator of value rows) for a workbook, values not formulas."""
    if src.suffix.lower() == ".xls":
        import xlrd  # base dependency; convert.py reads .xls with it too

        # logfile: xlrd narrates parse trivia straight to stdout otherwise.
        book = xlrd.open_workbook(str(src), logfile=io.StringIO())
        name = _pick_sheet(book.sheet_names(), sheet)
        sh = book.sheet_by_name(name)

        def cell(c):
            # Dates are bare floats in BIFF; only the cell type says so.
            if c.ctype == xlrd.XL_CELL_DATE:
                return xlrd.xldate_as_datetime(c.value, book.datemode)
            return c.value

        return name, ([cell(c) for c in sh.row(r)] for r in range(sh.nrows))
    from openpyxl import load_workbook

    wb = load_workbook(src, read_only=True, data_only=True)
    name = _pick_sheet(wb.sheetnames, sheet)
    return name, (list(r) for r in wb[name].iter_rows(values_only=True))


def _sheet_to_csv(src: Path, sheet: str, workdir: Path) -> tuple[str, Path]:
    """Extract one sheet to a CSV under `workdir`; (sheet name, csv path).

    A CSV round-trip on purpose: datetimes go out as ISO strings and numbers as
    their repr, both of which the full-file sniff recovers unambiguously — and
    typed values that openpyxl/xlrd already resolved (a formula's cached value,
    a BIFF date float) arrive as data, not artifacts.
    """
    import csv

    name, rows = _sheet_rows(src, sheet)
    out = workdir / "sheet.csv"
    with out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        for row in rows:
            writer.writerow(
                "" if v is None else v.isoformat() if hasattr(v, "isoformat") else v
                for v in row
            )
    return name, out


def _fit_rows(rows: list[list[Any]]) -> list[list[Any]]:
    """The longest prefix of `rows` that serializes under the byte cap."""
    size = 0
    for i, row in enumerate(rows):
        size += len(json.dumps(row, ensure_ascii=False, default=str)) + 2
        if size > _PAYLOAD_BYTE_CAP:
            return rows[:i]
    return rows


class QueryTableArgs(BaseModel):
    path: str = Field(
        description=(
            "Path to the data file to query "
            "(.csv/.tsv/.parquet/.json, or .xlsx/.xls for one sheet at a time)"
        ),
    )
    sql: str = Field(
        description=(
            "A single read-only SELECT. The file is bound to the table name `t` "
            "— e.g. SELECT region, avg(score) FROM t GROUP BY 1. "
            "When the columns are unknown, call with `SUMMARIZE t` first: one "
            "call returns every column's type, min/max, distinct count and "
            "null share, without spending rows."
        ),
    )
    limit: int = Field(
        default=200,
        description="Max rows returned; the reply flags whether it truncated",
    )
    sheet: str = Field(
        default="",
        description=(
            "Excel files only: which sheet to query. Omit for single-sheet "
            "workbooks; a multi-sheet workbook rejects the call and lists its "
            "sheet names."
        ),
    )


@tool(QueryTableArgs, cls="atomic")
def silica_query_table(
    path: str, sql: str, limit: int = 200, sheet: str = ""
) -> dict[str, Any]:
    """Answers a question about a data file (.csv/.tsv/.parquet/.json, Excel
    via sheet=) by running SQL over it — the aggregation path semantic search
    cannot do: sums, group-by, ranking, ranges. Queried in place, bound to
    table `t`. Read-only: one SELECT per call. `SUMMARIZE t` is the cheap
    first call when columns are unknown. Replies carry the bound schema; a
    numeric-looking VARCHAR column holds non-numbers — count what try_cast
    loses before trusting an aggregate.
    """
    src = Path(path).expanduser()
    # Vault first, cwd second — mirrors convert._resolve_input. The profile
    # note prints its query example repo-relative (machine-portable), so the
    # example must resolve from wherever the server happens to be running.
    if not src.is_absolute():
        from silica.config import CONFIG
        vault = (CONFIG.vault_path or "").strip()
        if vault and (Path(vault) / src).exists():
            src = Path(vault) / src
    try:
        src = src.resolve(strict=True)
    except OSError as e:
        raise ValueError(f"cannot read {path}: {e}") from e
    if not src.is_file():
        raise ValueError(f"{src} is not a file")
    suffix = src.suffix.lower()
    if suffix not in (*READABLE_EXTS, *SHEET_EXTS):
        raise ValueError(
            f"{suffix or 'no extension'} is not a tabular format — "
            f"expected one of {', '.join((*READABLE_EXTS, *SHEET_EXTS))}"
        )
    if sheet and suffix not in SHEET_EXTS:
        raise ValueError(f"sheet= applies to Excel files only, not {suffix}")

    inner = _select_only(sql)
    sheet_name = ""
    # The temp dir is the confinement boundary for the sheet path: the workbook
    # is read in-process BEFORE any model SQL runs, so the SQL can only see the
    # one extracted sheet.
    with tempfile.TemporaryDirectory() as tmp:
        if suffix in SHEET_EXTS:
            sheet_name, bound = _sheet_to_csv(src, sheet, Path(tmp))
        elif suffix in _SNIFFED_EXTS:
            bound = utf8_source(src, Path(tmp))
        else:
            bound = src
        con = _connect(bound.parent)
        try:
            _bind_source(con, bound)
            schema = dict(
                (row[0], row[1]) for row in con.execute("DESCRIBE t").fetchall()
            )
            # limit+1 so truncation is observed rather than guessed at.
            rows = con.execute(f"SELECT * FROM ({inner}) LIMIT {int(limit) + 1}").fetchall()
            columns = [d[0] for d in con.description]
        except Exception as e:
            raise ValueError(f"query failed: {type(e).__name__}: {e}") from e
        finally:
            con.close()

    over_limit = len(rows) > limit
    within_limit = [list(r) for r in rows[:limit]]
    kept = _fit_rows(within_limit)
    truncated = over_limit or len(kept) < len(within_limit)
    if over_limit:
        note = f"truncated at limit={limit}; aggregate or raise limit"
    elif truncated:
        note = (
            f"truncated at ~{_PAYLOAD_BYTE_CAP // 1000} KB of rows; "
            "aggregate or select fewer columns instead of scanning"
        )
    else:
        note = ""
    return {
        "path": str(src),
        **({"sheet": sheet_name} if sheet_name else {}),
        # The bound table's schema on every reply: SQL written against guessed
        # column names or types is the lane's main failure mode.
        "schema": schema,
        "columns": columns,
        "rows": kept,
        "row_count": len(kept),
        # Never a silent cap: a truncated answer that reads as complete is how a
        # BI number comes out confidently wrong.
        "truncated": truncated,
        **({"note": note} if note else {}),
    }

# One file per call, so no joins across files. Binding a dict of path→name as
# t1..tn is the design for the day a real question needs two tables at once.


# The census deliberately drops .txt and .json from READABLE_EXTS: a vault's
# .txt is prose more often than a table and .json is config more often than
# records, so listing them would bury the real tables under false positives.
# Both stay queryable by explicit path through silica_query_table.
_CENSUS_EXTS = (".csv", ".tsv", ".parquet", ".ndjson", *SHEET_EXTS)
_CATALOG_FILE_CAP = 500  # files walked before the census itself truncates
# Head sample, not query_table's whole-file sniff: a catalog is orientation,
# and the full sniff on a hundred-file vault costs seconds per call, while a
# head-typed column corrupts nothing here — query_table re-sniffs in full.
_CATALOG_SNIFF_ROWS = 2048


def _describe_table(src: Path) -> dict[str, Any]:
    """One census entry body: {"columns": …} | {"sheets": …} | {"error": …}.

    The census must finish: an unreadable file becomes a named error entry,
    never a dropped path or an aborted catalog.
    """
    import duckdb

    if src.suffix.lower() in SHEET_EXTS:
        # Sheets only, no per-sheet schema: opening every sheet of every
        # workbook is a query_table call's worth of work per file.
        try:
            if src.suffix.lower() == ".xls":
                import xlrd

                book = xlrd.open_workbook(str(src), logfile=io.StringIO())
                return {"sheets": book.sheet_names()}
            from openpyxl import load_workbook

            return {"sheets": load_workbook(src, read_only=True).sheetnames}
        except Exception as e:
            return {"error": f"{type(e).__name__}: {e}"}

    def describe(path: Path) -> dict[str, str]:
        quoted = str(path).replace("'", "''")
        rel = (
            f"read_csv('{quoted}', sample_size={_CATALOG_SNIFF_ROWS})"
            if path.suffix.lower() in _SNIFFED_EXTS
            else f"'{quoted}'"
        )
        # A bare connection, not _connect's sandbox: the SQL here is ours (a
        # DESCRIBE over a resolved vault path), never model-authored.
        con = duckdb.connect()
        try:
            rows = con.execute(f"DESCRIBE SELECT * FROM {rel}").fetchall()
        finally:
            con.close()
        return {r[0]: r[1] for r in rows}

    try:
        return {"columns": describe(src)}
    except Exception as first:
        if src.suffix.lower() in _SNIFFED_EXTS:
            # Retry through the encoding ladder before giving up: public-data
            # exports are routinely cp1252 (4 of 26 CSVs on a real vault) and
            # DuckDB refuses them outright.
            try:
                with tempfile.TemporaryDirectory() as tmp:
                    return {"columns": describe(utf8_source(src, Path(tmp)))}
            except Exception:
                pass
        return {"error": f"{type(first).__name__}: {first}"}


class TablesArgs(BaseModel):
    folder: str = Field(
        default="",
        description="Vault-relative folder to scope the census; empty = the whole vault",
    )
    column: str = Field(
        default="",
        description=(
            "Case-insensitive substring filter: only tables with a matching "
            "column name return, with the matches named"
        ),
    )
    limit: int = Field(default=50, description="Max tables described in the reply")


@tool(TablesArgs, cls="atomic")
def silica_tables(folder: str = "", column: str = "", limit: int = 50) -> dict[str, Any]:
    """Census of the vault's tabular files (.csv/.tsv/.parquet/.ndjson, Excel):
    path, size and column schema per file, filterable by column name — the
    orientation call before silica_query_table. column= answers "which table
    holds NEET?" in one call instead of head-reading every file. Excel entries
    list their sheets (a sheet's schema costs one query_table call). Schemas
    come from a head sample; the types query_table replies with (whole-file
    sniff) are the ones to trust."""
    import os

    from silica.config import CONFIG
    from silica.kernel.recall.paths import ignore_matcher

    vault = (CONFIG.vault_path or "").strip()
    if not vault:
        raise ValueError("no vault configured")
    root = Path(vault).resolve()
    base = (root / folder).resolve() if folder.strip() else root
    try:
        base.relative_to(root)
    except ValueError:
        raise ValueError(f"{folder!r} escapes the vault") from None
    if not base.is_dir():
        raise ValueError(f"{folder or str(root)!r} is not a folder in the vault")

    ignored = ignore_matcher(root)
    found: list[Path] = []
    for dirpath, dirnames, filenames in os.walk(base):
        # Same pruning as the note walks: dot-dirs, NOISE_DIRS + .silicaignore.
        # Images/ too — conversion output, not data.
        dirnames[:] = sorted(
            d for d in dirnames
            if not d.startswith(".") and not ignored(d) and d != "Images"
        )
        for name in sorted(filenames):
            if not name.startswith(".") and Path(name).suffix.lower() in _CENSUS_EXTS:
                found.append(Path(dirpath) / name)
        if len(found) > _CATALOG_FILE_CAP:
            break
    walk_truncated = len(found) > _CATALOG_FILE_CAP
    found = found[:_CATALOG_FILE_CAP]

    needle = column.strip().casefold()
    tables: list[dict[str, Any]] = []
    unsearchable = 0
    scan_stopped = False
    for src in found:
        if len(tables) >= max(1, limit):
            # Files past this point were never described: truncated, not absent.
            scan_stopped = True
            break
        body = _describe_table(src)
        if needle:
            cols = body.get("columns")
            if cols is None:
                unsearchable += 1  # sheets/errors carry no columns to match
                continue
            matched = sorted(c for c in cols if needle in c.casefold())
            if not matched:
                continue
            body = {**body, "matched": matched}
        tables.append({
            "path": src.relative_to(root).as_posix(),
            "size_bytes": src.stat().st_size,
            **body,
        })

    # Same byte discipline as query_table: 50 wide schemas can outweigh 200 rows.
    kept, size = [], 0
    for entry in tables:
        size += len(json.dumps(entry, ensure_ascii=False, default=str)) + 2
        if size > _PAYLOAD_BYTE_CAP:
            break
        kept.append(entry)

    truncated = walk_truncated or scan_stopped or len(kept) < len(tables)
    notes = []
    if walk_truncated:
        notes.append(f"census stopped at {_CATALOG_FILE_CAP} files; narrow with folder=")
    if scan_stopped:
        notes.append(f"stopped at limit={limit}; raise limit or narrow with folder=")
    if len(kept) < len(tables):
        notes.append(f"~{_PAYLOAD_BYTE_CAP // 1000} KB payload cap hit; narrow the census")
    if needle and unsearchable:
        notes.append(
            f"{unsearchable} file(s) had no readable columns to match "
            "(Excel workbooks or unreadable files)"
        )
    return {
        # The census size, before filter and caps: "3 of 143" reads correctly.
        "total": len(found),
        "returned": len(kept),
        "tables": kept,
        "truncated": truncated,
        **({"note": "; ".join(notes)} if notes else {}),
    }
